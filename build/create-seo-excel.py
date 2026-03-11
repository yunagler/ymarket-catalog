"""
Y Market - SEO Products Excel Generator
Creates an Excel file with all products grouped by category
for manual SEO data entry.
"""
import json, sys
sys.stdout.reconfigure(encoding='utf-8')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

data = json.load(open('data/products.json', encoding='utf-8'))
items = data['items']

wb = Workbook()

# Sort by category then name
items_sorted = sorted(items, key=lambda x: (x.get('categoryName',''), x.get('name','')))

# Group by category
cats = {}
for item in items_sorted:
    cat = item.get('categoryName', 'אחר')
    if cat not in cats:
        cats[cat] = []
    cats[cat].append(item)

# Styles
header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='1B3A5C', end_color='1B3A5C', fill_type='solid')
cat_font = Font(name='Arial', bold=True, size=12, color='1B3A5C')
cat_fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
existing_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)
wrap_align = Alignment(wrap_text=True, vertical='top', horizontal='right')
header_align = Alignment(wrap_text=True, vertical='center', horizontal='center')

# Headers
headers = [
    ('מס"ד', 6),
    ('מזהה', 8),
    ('שם המוצר', 35),
    ('קטגוריה', 22),
    ('יחידה/אריזה', 15),
    ('מחיר ₪', 10),
    # SEO fields for user to fill
    ('כותרת SEO (H1) - עד 70 תווים', 40),
    ('תיאור מטא - 120-160 תווים (מופיע בגוגל)', 50),
    ('תיאור מוצר מפורט (2-4 משפטים)', 55),
    ('חומר / הרכב', 20),
    ('מפרט 1 - שם', 15),
    ('מפרט 1 - ערך', 20),
    ('מפרט 2 - שם', 15),
    ('מפרט 2 - ערך', 20),
    ('מפרט 3 - שם', 15),
    ('מפרט 3 - ערך', 20),
    ('מפרט 4 - שם', 15),
    ('מפרט 4 - ערך', 20),
    ('מילות חיפוש (מופרדות בפסיק)', 35),
    ('CTA כמות - כותרת', 25),
    ('CTA כמות - טקסט', 35),
    ('הערות', 25),
]

ws = wb.active
ws.title = 'מוצרים SEO'
ws.sheet_view.rightToLeft = True

# Write headers
for col_idx, (header, width) in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Freeze first row
ws.freeze_panes = 'A2'

row = 2
num = 1
for cat_name, cat_items in cats.items():
    # Category header row
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = cat_fill
        cell.border = thin_border
    ws.cell(row=row, column=3, value=f'{cat_name} ({len(cat_items)} פריטים)')
    ws.cell(row=row, column=3).font = cat_font
    ws.cell(row=row, column=3).fill = cat_fill
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    row += 1

    for item in cat_items:
        seo = item.get('seo', {})
        specs = seo.get('specs', [])
        bulk = seo.get('bulkCta', {})

        # Basic data (pre-filled, gray background)
        for col_idx in range(1, 7):
            ws.cell(row=row, column=col_idx).fill = existing_fill

        ws.cell(row=row, column=1, value=num).border = thin_border
        ws.cell(row=row, column=2, value=item.get('id','')).border = thin_border
        ws.cell(row=row, column=3, value=item.get('name','')).border = thin_border
        ws.cell(row=row, column=4, value=cat_name).border = thin_border
        ws.cell(row=row, column=5, value=item.get('unit','')).border = thin_border
        ws.cell(row=row, column=6, value=item.get('saleNis','')).border = thin_border

        # SEO fields - pre-fill if seo.specs exists (manual data)
        ws.cell(row=row, column=7, value=seo.get('h1', '')).border = thin_border
        ws.cell(row=row, column=8, value=seo.get('metaDesc', '')).border = thin_border
        ws.cell(row=row, column=9, value=item.get('description', '')).border = thin_border

        # Material
        ws.cell(row=row, column=10, value='').border = thin_border

        # Specs (up to 4)
        for i in range(4):
            if i < len(specs):
                ws.cell(row=row, column=11 + i*2, value=specs[i].get('label','')).border = thin_border
                ws.cell(row=row, column=12 + i*2, value=specs[i].get('value','')).border = thin_border
            else:
                ws.cell(row=row, column=11 + i*2, value='').border = thin_border
                ws.cell(row=row, column=12 + i*2, value='').border = thin_border

        # Search tags
        ws.cell(row=row, column=19, value=', '.join(item.get('searchTags', []))).border = thin_border

        # Bulk CTA
        ws.cell(row=row, column=20, value=bulk.get('title', '')).border = thin_border
        ws.cell(row=row, column=21, value=bulk.get('text', '')).border = thin_border

        # Notes
        ws.cell(row=row, column=22, value='').border = thin_border

        # Alignment for all cells
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row, column=col_idx).alignment = wrap_align

        num += 1
        row += 1

# ── Instructions sheet ──
ws2 = wb.create_sheet('הוראות מילוי')
ws2.sheet_view.rightToLeft = True

instr = [
    (1, 'הוראות מילוי טבלת SEO למוצרים - וואי מרקט', True, 14, '1B3A5C'),
    (3, 'עמודות אפורות (1-6) = נתונים קיימים מהמערכת. לא לשנות!', True, 11, 'CC0000'),
    (4, 'עמודות לבנות (7-22) = למלא ידנית', True, 11, '000000'),
    (5, '', False, 10, '000000'),
    (6, '── פירוט העמודות ──', True, 12, '1B3A5C'),
    (7, '', False, 10, '000000'),
    (8, 'כותרת SEO (H1) [עמודה 7]', True, 11, '000000'),
    (9, '   כותרת שתופיע כ-H1 בדף המוצר. צריכה להכיל מילות חיפוש רלוונטיות.', False, 10, '444444'),
    (10, '   עד 70 תווים. דוגמה: "גליל ניילון נצמד 6 קילו – עובי מקצועי לעיטוף מגשי בשר"', False, 10, '16a34a'),
    (11, '', False, 10, '000000'),
    (12, 'תיאור מטא [עמודה 8]', True, 11, '000000'),
    (13, '   התיאור הקצר שמופיע בתוצאות גוגל. 120-160 תווים. כולל מילות מפתח + קריאה לפעולה.', False, 10, '444444'),
    (14, '   דוגמה: "ניילון נצמד תעשייתי 6 ק״ג. כושר מתיחה גבוה, שקיפות קריסטלית. מותאם לקצביות. מחירי סיטונאות + משלוח."', False, 10, '16a34a'),
    (15, '', False, 10, '000000'),
    (16, 'תיאור מוצר מפורט [עמודה 9]', True, 11, '000000'),
    (17, '   התיאור המלא שמופיע בדף המוצר. 2-4 משפטים. מה המוצר, למה הוא טוב, למי מתאים.', False, 10, '444444'),
    (18, '   חשוב: טקסט ייחודי! לא העתקה מאתרים אחרים.', False, 10, 'CC0000'),
    (19, '', False, 10, '000000'),
    (20, 'חומר / הרכב [עמודה 10]', True, 11, '000000'),
    (21, '   מה החומר של המוצר. דוגמה: LLDPE, פוליפרופילן PP, נתרן היפוכלוריט 3.5%', False, 10, '444444'),
    (22, '', False, 10, '000000'),
    (23, 'מפרט 1-4 [עמודות 11-18]', True, 11, '000000'),
    (24, '   עד 4 שורות מפרט טכני. כל שורה = שם שדה + ערך.', False, 10, '444444'),
    (25, '   דוגמה: שם="משקל גליל" ערך="6 ק״ג" | שם="מאושר למגע מזון" ערך="כן"', False, 10, '16a34a'),
    (26, '', False, 10, '000000'),
    (27, 'מילות חיפוש [עמודה 19]', True, 11, '000000'),
    (28, '   מילים שאנשים מחפשים בגוגל, מופרדות בפסיק. מילים נרדפות, שמות חלופיים.', False, 10, '444444'),
    (29, '   דוגמה: ניילון נצמד, stretch film, ניילון עיטוף, ניילון מזון, ניילון למכונה', False, 10, '16a34a'),
    (30, '', False, 10, '000000'),
    (31, 'CTA כמות [עמודות 20-21]', True, 11, '000000'),
    (32, '   כותרת וטקסט לכפתור "צריכים כמות גדולה?" שמפנה לוואטסאפ. אופציונלי.', False, 10, '444444'),
    (33, '', False, 10, '000000'),
    (34, '── טיפים ──', True, 12, '1B3A5C'),
    (35, '', False, 10, '000000'),
    (36, 'המוצר "נילון נצמד מכונה 6 ק״ג" כבר ממולא = דוגמה מושלמת!', True, 11, '16a34a'),
    (37, 'אפשר למלא רק עמודות 7-9 (הכי חשובות) ולהשאיר את השאר ריק.', False, 10, '444444'),
    (38, 'אם יש מוצרים דומים - אפשר להעתיק תיאור ולשנות פרטים.', False, 10, '444444'),
    (39, 'אחרי שתסיים - שלח לי את הקובץ ואני אעדכן את האתר אוטומטית!', True, 11, '1B3A5C'),
]

for row_num, text, bold, size, color in instr:
    cell = ws2.cell(row=row_num, column=1, value=text)
    cell.font = Font(name='Arial', bold=bold, size=size, color=color)
    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='right')

ws2.column_dimensions['A'].width = 100

# Auto-filter on main sheet
ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{row-1}'

output = r'C:\Users\DELL\Desktop\ymarket_seo_products.xlsx'
wb.save(output)
print(f'Saved: {output}')
print(f'Products: {num-1}')
print(f'Categories: {len(cats)}')
